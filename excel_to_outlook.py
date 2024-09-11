from typing import List
import win32com.client
import openpyxl
import datetime

import argparse
from pathlib import Path

class Appointment:
    #enums: https://learn.microsoft.com/en-us/office/vba/api/outlook

    def __init__(self, subject: str, start_date: datetime.datetime, start_time: datetime.time, duration:datetime.time, location:str, recipients: List[str], body: str):
        self.subject = subject

        date_and_time = datetime.datetime.combine(start_date, start_time)
        self.start = date_and_time.strftime("%Y-%m-%d %H:%M")
        
        self.duration = duration.hour * 60 + duration.minute #in minutes
        self.location = "tbd" if location is None else location

        self.recipients = recipients
        self.body = "" if body is None else body

    def create_outlook(self):
        outlook = win32com.client.Dispatch("Outlook.Application")
        appt = outlook.CreateItem(1) # 1 = olAppointmentItem

        if len(self.recipients) > 0:
            appt.MeetingStatus = 1
            for recipient in self.recipients:
                attendee = appt.Recipients.Add(recipient)
                attendee.Type = 1 # 1 = olRequired

        appt.Subject = self.subject
        appt.Start = self.start
        appt.Duration = self.duration
        appt.Body = self.body
        appt.Location = self.location
        appt.Organizer = "Excel to Outlook"       

        appt.BusyStatus = 2 # 2 = olBusy
        appt.ReminderSet = True
        appt.ReminderMinutesBeforeStart = 15

        appt.Save()
        appt.Send()

def data_correct(raw_data: dict):
    mandatory_fields = ["Betreff", "Start Datum", "Start Uhrzeit", "Dauer"]
    for field in mandatory_fields:
        if raw_data[field] is None:
            return False

    return True

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Reads an excel file and creates appointments in Outlook")
    parser.add_argument("path", help="The excel file to read from")
    args = parser.parse_args()

    path = Path(args.path)
    if not path.exists():
        print(f"File {path} does not exist")
        exit()

    dataframe = openpyxl.load_workbook(path)
    active_sheet = dataframe.active
 
    for row in range(1, active_sheet.max_row): #0 is the header
        raw_data = {}
        for col in active_sheet.iter_cols(1, active_sheet.max_column):
            raw_data[col[0].value] = col[row].value
 
        recipients_str: str = raw_data["Empfänger"]
        if recipients_str is None:
            recipients = []
        else:
            recipients = [x.strip() for x in recipients_str.split(";")]
 
        if data_correct(raw_data):
            appointment = Appointment(raw_data["Betreff"], raw_data["Start Datum"], raw_data["Start Uhrzeit"], raw_data["Dauer"], raw_data["Ort"], recipients, raw_data["Inhalt"])
            appointment.create_outlook()
        
    print("Done!")